from flask import Flask, jsonify, render_template, request, make_response, session, redirect, url_for, flash
from flask_socketio import SocketIO, emit, join_room, leave_room
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import random
import time
from datetime import datetime, date, UTC, timedelta
import os
from database import (get_db, init_db, get_user_by_username, create_user, get_user_by_id,
                      get_user_settings, update_user_settings, reset_user_settings,
                      create_verification_token, verify_email_token, cleanup_expired_tokens,
                      get_user_by_email, create_password_reset_token, verify_reset_token,
                      reset_password, cleanup_expired_reset_tokens, log_login, get_login_history,
                      is_admin, set_user_role, get_all_users, get_admin_stats,
                      log_audit, get_audit_logs, cleanup_old_audit_logs, cleanup_old_data,
                      cleanup_old_login_history, delete_user_with_audit, get_database_info,
                      init_onboarding, get_onboarding_status, update_onboarding_step, 
                      mark_feature_as_seen, complete_onboarding, start_tour, complete_tour,
                      create_scheduled_export, get_user_scheduled_exports, get_due_exports,
                      update_export_timestamp, delete_scheduled_export,
                      get_user_thresholds, update_user_thresholds, check_threshold_status,
                      grant_permission, revoke_permission, has_permission, get_user_permissions, get_users_with_permission,
                      import_csv_readings, get_csv_import_stats,
                      create_sensor, get_user_sensors, get_sensor_by_id, update_sensor, delete_sensor,
                      get_active_sensors, update_sensor_availability, update_sensor_last_read,
                      log_sensor_reading, get_sensor_readings, get_sensor_latest_reading,
                      update_sensor_thresholds, get_sensor_thresholds, check_sensor_threshold_status)
import json
from flask import send_file
try:
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception as e:
    HTML = None
    WEASYPRINT_AVAILABLE = False
    # WeasyPrint requires system libraries (GTK) - PDF export will be unavailable
    # This is normal on Windows and can be safely ignored
import io
import threading
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import csv
import secrets

from fake_co2 import generate_co2, generate_co2_data, save_reading, reset_state, set_scenario, get_scenario_info, set_paused
from database import cleanup_old_data
from advanced_features import (AdvancedAnalytics, CollaborationManager, 
                               PerformanceOptimizer, VisualizationEngine)
from advanced_features_routes import register_advanced_features



app = Flask(__name__)
app.config['SECRET_KEY'] = 'morpheus-co2-secret-key'

# Initialize rate limiter - disabled for live data polling
# Uncomment and configure if you need rate limiting in production
# limiter = Limiter(
#     app=app,
#     key_func=get_remote_address,
#     default_limits=["500 per day", "150 per hour"],
#     storage_uri="memory://"
# )

# Create a dummy limiter for compatibility with existing routes
class DummyLimiter:
    def limit(self, *args, **kwargs):
        def decorator(f):
            return f
        return decorator
    
    def exempt(self, f):
        return f

limiter = DummyLimiter()

# Email configuration (using development/testing settings)
# In production, use environment variables for credentials
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS', True)
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER', 'noreply@morpheus-co2.local')

socketio = SocketIO(
    app, 
    cors_allowed_origins="*",
    async_mode='threading',
    logger=False,
    engineio_logger=False,
    ping_timeout=60,
    ping_interval=25
)

init_db()

# Register advanced features routes
register_advanced_features(app, limiter)

# Register Jinja2 globals for templates
app.jinja_env.globals['is_admin'] = is_admin

# ================================================================================
#                    CONTEXT PROCESSOR
# ================================================================================

@app.context_processor
def inject_user_context():
    """Inject user context variables into all templates"""
    user_is_admin = False
    if 'user_id' in session:
        user_is_admin = is_admin(session['user_id'])
    
    return {
        'current_user_is_admin': user_is_admin,
        'is_logged_in': 'user_id' in session
    }

# ================================================================================
#                    SECURITY HEADERS & MIDDLEWARE
# ================================================================================

@app.after_request
def add_security_headers(response):
    """Add security headers to all responses"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' cdn.jsdelivr.net cdn.socket.io; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:;"
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

def send_verification_email(email, username, token):
    """Send email verification link"""
    try:
        from flask_mail import Mail, Message
        mail = Mail(app)
        
        verify_url = url_for('verify_email', token=token, _external=True)
        subject = "Verify your Morpheus CO₂ Account"
        
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <div style="background: linear-gradient(135deg, #3dd98f 0%, #4db8ff 100%); padding: 20px; border-radius: 10px; color: white; text-align: center; margin-bottom: 20px;">
                        <h1 style="margin: 0;">🌍 Morpheus CO₂ Monitor</h1>
                    </div>
                    
                    <h2 style="color: #3dd98f;">Welcome, {username}!</h2>
                    <p>Thank you for registering with Morpheus. Please verify your email address to complete your account setup.</p>
                    
                    <div style="background: #f5f5f5; padding: 20px; border-radius: 8px; text-align: center; margin: 20px 0;">
                        <p style="margin-bottom: 15px;">Click the button below to verify your email:</p>
                        <a href="{verify_url}" style="background: linear-gradient(135deg, #3dd98f 0%, #4db8ff 100%); color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">Verify Email</a>
                    </div>
                    
                    <p style="font-size: 12px; color: #666;">Or copy and paste this link in your browser:</p>
                    <p style="font-size: 11px; color: #666; word-break: break-all; background: #f0f0f0; padding: 10px; border-radius: 4px;">{verify_url}</p>
                    
                    <p style="font-size: 12px; color: #999; margin-top: 20px; border-top: 1px solid #ddd; padding-top: 20px;">
                        This link will expire in 24 hours. If you didn't create this account, please ignore this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        msg = Message(subject=subject, recipients=[email], html=html_body)
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        return False

def send_password_reset_email(email, username, token):
    """Send password reset email"""
    try:
        from flask_mail import Mail, Message
        mail = Mail(app)
        
        reset_url = url_for('reset_password_page', token=token, _external=True)
        subject = "Reset your Morpheus CO₂ Password"
        
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <div style="background: linear-gradient(135deg, #3dd98f 0%, #4db8ff 100%); padding: 20px; border-radius: 10px; color: white; text-align: center; margin-bottom: 20px;">
                        <h1 style="margin: 0;">🌍 Morpheus CO₂ Monitor</h1>
                    </div>
                    
                    <h2 style="color: #3dd98f;">Password Reset Request</h2>
                    <p>Hi {username},</p>
                    <p>You requested to reset your password. Click the button below to proceed:</p>
                    
                    <div style="background: #f5f5f5; padding: 20px; border-radius: 8px; text-align: center; margin: 20px 0;">
                        <p style="margin-bottom: 15px;">Click the button below to reset your password:</p>
                        <a href="{reset_url}" style="background: linear-gradient(135deg, #3dd98f 0%, #4db8ff 100%); color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">Reset Password</a>
                    </div>
                    
                    <p style="font-size: 12px; color: #666;">Or copy and paste this link in your browser:</p>
                    <p style="font-size: 11px; color: #666; word-break: break-all; background: #f0f0f0; padding: 10px; border-radius: 4px;">{reset_url}</p>
                    
                    <p style="font-size: 12px; color: #999; margin-top: 20px; border-top: 1px solid #ddd; padding-top: 20px;">
                        This link will expire in 1 hour. If you didn't request this, please ignore this email and your password will remain unchanged.
                    </p>
                </div>
            </body>
        </html>
        """
        
        msg = Message(subject=subject, recipients=[email], html=html_body)
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send password reset email: {str(e)}")
        return False

def load_settings():
    db = get_db()
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    db.close()

    settings = DEFAULT_SETTINGS.copy()
    for r in rows:
        settings[r["key"]] = json.loads(r["value"])

    return settings

print("=" * 50)
print(f"Current directory: {os.getcwd()}")
print(f"Template folder exists: {os.path.exists('templates')}")
if os.path.exists('templates'):
    print(f"Files in templates/: {os.listdir('templates')}")
print("=" * 50)

DEFAULT_SETTINGS = {
    "analysis_running": True,
    "good_threshold": 800,
    "warning_threshold": 1000,
    "bad_threshold": 1200,
    "critical_threshold": 1200,
    "alert_threshold": 1400,
    "realistic_mode": True,
    "update_speed": 1,
    "overview_update_speed": 5,
    "simulate_live": False,
}

# Centralized source aliases to keep simulation data isolated from live UI
REAL_SOURCES = ("live", "sensor", "real", "live_real")
SIM_SOURCES = ("sim",)
IMPORT_SOURCES = ("import",)


def resolve_source_param(*, default="live", allow_sim=False, allow_import=False):
    """Normalize ?source= query to an allowed bucket (live/sim/import)."""
    raw = (request.args.get("source") or default)
    normalized_map = {
        "live": "live",
        "real": "live",
        "hardware": "live",
        "simulation": "sim",
        "sim": "sim",
        "import": "import",
        "csv": "import",
    }
    normalized = normalized_map.get(str(raw).lower(), "live")

    if normalized == "sim" and not allow_sim:
        normalized = "live"
    if normalized == "import" and not allow_import:
        normalized = "live"

    return normalized


def build_source_filter(db_source):
    """Return SQL clause and params for the chosen data source bucket."""
    if db_source == "sim":
        return "source = ?", SIM_SOURCES
    if db_source == "import":
        return "source = ?", IMPORT_SOURCES
    return f"source IN ({','.join('?' * len(REAL_SOURCES))})", REAL_SOURCES

def save_settings(data):
    db = get_db()
    for k, v in data.items():
        db.execute(
            "REPLACE INTO settings (key, value) VALUES (?, ?)",
            (k, json.dumps(v))
        )
    db.commit()
    db.close()

# ================================================================================
#                        AUTHENTICATION DECORATOR
# ================================================================================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page', next=request.url))
        if not is_admin(session['user_id']):
            return render_template("error.html", 
                error="Access Denied",
                message="You need administrator privileges to access this page."), 403
        return f(*args, **kwargs)
    return decorated_function

def permission_required(permission):
    """Decorator to check if user has a specific permission"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login_page', next=request.url))
            if not has_permission(session['user_id'], permission):
                return render_template("error.html",
                    error="Access Denied",
                    message=f"You do not have the '{permission}' permission."), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ================================================================================
#                        AUTHENTICATION ROUTES
# ================================================================================

@app.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute")
def login_page():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        remember_me = request.form.get("remember_me") == "on"
        
        if not username or not password:
            return render_template("login.html", error="Nom d'utilisateur et mot de passe requis")
        
        user = get_user_by_username(username)
        
        if user and check_password_hash(user['password_hash'], password):
            # Login successful
            session['user_id'] = user['id']
            session['username'] = user['username']
            
            # Log successful login
            ip_address = request.remote_addr
            user_agent = request.headers.get('User-Agent', 'Unknown')
            log_login(user['id'], ip_address, user_agent, success=True)
            
            # Handle "Remember Me" - extend session duration
            if remember_me:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=30)
            
            # Check if user has completed onboarding
            onboarding = get_onboarding_status(user['id'])
            if not onboarding or not onboarding.get('completed'):
                return redirect(url_for('onboarding_page'))
            
            next_page = request.args.get('next')
            if next_page and next_page.startswith('/'):
                return redirect(next_page)
            return redirect(url_for('index'))
        else:
            return render_template("login.html", error="Identifiants invalides")
    
    return render_template("login.html")

@app.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("3 per minute")
def forgot_password_page():
    """Request password reset"""
    if request.method == "POST":
        email = request.form.get("email")
        
        if not email:
            return render_template("forgot_password.html", error="Veuillez entrer votre email")
        
        user = get_user_by_email(email)
        
        if user:
            # Generate reset token
            token = secrets.token_urlsafe(32)
            expires_at = datetime.now(UTC) + timedelta(hours=1)
            create_password_reset_token(user['id'], token, expires_at)
            
            # Send reset email
            email_sent = send_password_reset_email(email, user['username'], token)
            
            if email_sent:
                return render_template("forgot_password.html",
                    success=True,
                    message=f"Vérifiez votre email ({email}) pour obtenir le lien de réinitialisation du mot de passe.")
            else:
                # Email service not configured, still show success for security
                return render_template("forgot_password.html",
                    success=True,
                    message="Si ce compte existe, un email de réinitialisation a été envoyé.")
        else:
            # Don't reveal if email exists (security best practice)
            return render_template("forgot_password.html",
                success=True,
                message="Si ce compte existe, un email de réinitialisation a été envoyé.")
    
    return render_template("forgot_password.html")

@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password_page(token):
    """Reset password with token"""
    user_id = verify_reset_token(token)
    
    if not user_id:
        return render_template("reset_password.html", 
            error="Lien de réinitialisation invalide ou expiré.", 
            valid_token=False), 400
    
    if request.method == "POST":
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")
        
        # Validation
        if not new_password or not confirm_password:
            return render_template("reset_password.html", error="Tous les champs sont requis", valid_token=True)
        
        if len(new_password) < 6:
            return render_template("reset_password.html", 
                error="Le mot de passe doit contenir au moins 6 caractères", 
                valid_token=True)
        
        if new_password != confirm_password:
            return render_template("reset_password.html", 
                error="Les mots de passe ne correspondent pas", 
                valid_token=True)
        
        # Reset password
        new_password_hash = generate_password_hash(new_password)
        reset_password(user_id, new_password_hash, token)
        
        return render_template("reset_password.html", 
            success=True, 
            message="Votre mot de passe a été réinitialisé avec succès! Vous pouvez maintenant vous connecter.")
    
    return render_template("reset_password.html", valid_token=True)

@app.route("/register", methods=["GET", "POST"])
@limiter.limit("3 per minute")
def register_page():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        # Validation
        if not all([username, email, password, confirm_password]):
            return render_template("register.html", error="Tous les champs sont requis")
        
        if len(username) < 3:
            return render_template("register.html", error="Le nom d'utilisateur doit contenir au moins 3 caractères")
        
        if len(password) < 6:
            return render_template("register.html", error="Le mot de passe doit contenir au moins 6 caractères")
        
        if password != confirm_password:
            return render_template("register.html", error="Les mots de passe ne correspondent pas")
        
        # Check if username or email already exists
        if get_user_by_username(username):
            return render_template("register.html", error="Ce nom d'utilisateur existe déjà")
        
        # Create user
        password_hash = generate_password_hash(password)
        user_id = create_user(username, email, password_hash)
        
        if not user_id:
            return render_template("register.html", error="Cet email existe déjà")
        
        # Generate verification token
        token = secrets.token_urlsafe(32)
        expires_at = datetime.now(UTC) + timedelta(hours=24)
        create_verification_token(user_id, token, expires_at)
        
        # Send verification email
        email_sent = send_verification_email(email, username, token)
        
        if email_sent:
            return render_template("register.html", 
                success=True,
                message=f"Inscription réussie! Vérifiez votre email ({email}) pour confirmer votre compte.")
        else:
            # Email service not configured, allow login without verification
            session['user_id'] = user_id
            session['username'] = username
            # Redirect to onboarding for new users
            return redirect(url_for('onboarding_page'))
    
    return render_template("register.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route("/verify/<token>")
def verify_email(token):
    """Verify user email with token"""
    user_id = verify_email_token(token)
    
    if user_id:
        user = get_user_by_id(user_id)
        if user:
            return render_template("email_verified.html", username=user['username'], success=True)
        else:
            return render_template("email_verified.html", 
                error="Lien de vérification invalide ou expiré. Veuillez vous réinscrire.",
                success=False), 400
    else:
        return render_template("email_verified.html", 
            error="Lien de vérification invalide ou expiré. Veuillez vous réinscrire.",
            success=False), 400

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user_id = session.get('user_id')
    user = get_user_by_id(user_id)
    user_settings = get_user_settings(user_id)
    login_history = get_login_history(user_id, limit=5)  # Get last 5 logins
    admin_stats = None
    admin_users = None
    
    # Check if user is admin to load admin dashboard data
    if is_admin(user_id):
        admin_stats = get_admin_stats()
        admin_users = get_all_users()
    
    # Note: CO₂ threshold settings are now handled in /settings page
    # Profile page is only for user info and admin dashboard
    if request.method == "POST":
        # This route no longer handles form submissions
        # CO₂ settings go to /api/settings
        pass
    
    return render_template("profile.html", 
                         user=user, 
                         settings=user_settings, 
                         login_history=login_history,
                         is_admin=is_admin(user_id),
                         admin_stats=admin_stats,
                         admin_users=admin_users)

@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user_id = session.get('user_id')
    user = get_user_by_id(user_id)
    
    if not user:
        session.clear()
        return redirect(url_for('login_page'))
    
    if request.method == "POST":
        current_password = request.form.get('current_password', "")
        new_password = request.form.get('new_password', "")
        confirm_password = request.form.get('confirm_password', "")
        
        error = None
        
        # Validation
        if not all([current_password, new_password, confirm_password]):
            error = "Tous les champs sont requis"
        elif not check_password_hash(user['password_hash'], current_password):
            error = "Mot de passe actuel incorrect"
        elif len(new_password) < 6:
            error = "Le nouveau mot de passe doit contenir au moins 6 caractères"
        elif new_password != confirm_password:
            error = "Les mots de passe ne correspondent pas"
        
        if error:
            return render_template("change_password.html", error=error)
        
        # Update password
        db = get_db()
        new_password_hash = generate_password_hash(new_password)
        db.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (new_password_hash, user_id)
        )
        db.commit()
        db.close()
        
        return render_template("change_password.html", success=True)
    
    return render_template("change_password.html", user=user)
    return render_template("profile.html", user=user, settings=user_settings)

# 1. ROOT ROUTE - DASHBOARD (MUST BE FIRST!)
@app.route("/")
@login_required
def index():
    return render_template("index.html")  # Overview/vue d'ensemble page

@app.route("/dashboard")
@login_required
def dashboard():
    """Unified landing page - shows admin dashboard for admins, user info for others"""
    user_id = session.get('user_id')
    
    if is_admin(user_id):
        # Show admin dashboard with stats
        admin_stats = get_admin_stats()
        admin_users = get_all_users()
        return render_template("dashboard.html", 
                             is_admin=True,
                             admin_stats=admin_stats,
                             admin_users=admin_users)
    else:
        # Show regular user dashboard with user-specific stats
        user = get_user_by_id(user_id)
        user_settings = get_user_settings(user_id)
        login_history = get_login_history(user_id, limit=10)
        
        # Calculate user statistics from database
        db = get_db()
        cursor = db.cursor()
        
        # Get today's reading stats
        cursor.execute("""
            SELECT 
                COUNT(*) as total_readings,
                AVG(ppm) as avg_ppm,
                MAX(ppm) as max_ppm,
                MIN(ppm) as min_ppm
            FROM readings
            WHERE DATE(timestamp) = CURRENT_DATE
        """)
        today_stats = cursor.fetchone()
        
        # Get this week's stats
        cursor.execute("""
            SELECT 
                COUNT(*) as total_readings,
                AVG(ppm) as avg_ppm
            FROM readings
            WHERE DATE(timestamp) >= DATE('now', '-7 days')
        """)
        week_stats = cursor.fetchone()
        
        # Get high CO2 events (above threshold)
        user_threshold = user_settings.get('co2_threshold', 800) if user_settings else 800
        cursor.execute("""
            SELECT COUNT(*) as bad_events
            FROM readings
            WHERE DATE(timestamp) = CURRENT_DATE AND ppm > ?
        """, (user_threshold,))
        bad_events = cursor.fetchone()
        
        db.close()
        
        return render_template("dashboard.html", 
                             is_admin=False,
                             user=user,
                             user_settings=user_settings,
                             login_history=login_history,
                             today_stats=today_stats,
                             week_stats=week_stats,
                             bad_events=bad_events)

@app.route("/live")
@login_required
def live_page():
    return render_template("live.html")  # Settings page

# 2. SETTINGS ROUTE
@app.route("/settings")
@login_required
def settings_page():
    return render_template("settings.html")  # Settings page

@app.route("/sensors")
@login_required
def sensors_page():
    """Sensor management and configuration page"""
    return render_template("sensors.html")

@app.route("/simulator")
@login_required
def simulator_page():
    """Simulator control page for testing scenarios - accessible to all users"""
    return render_template("simulator.html")

@app.route("/visualization")
@login_required
def visualization():
    """Advanced data visualization dashboard with CSV import"""
    return render_template("visualization.html")

@app.route("/features-hub")
@login_required
def features_hub():
    """Feature hub - Main landing page with all features"""
    return render_template("features-hub.html")

@app.route("/advanced-features")
@login_required
def advanced_features_page():
    """Advanced features dashboard - Analytics, Insights, Visualizations, Collaboration, Performance"""
    return render_template("advanced-features.html")

@app.route("/analytics")
@login_required
def analytics_feature():
    """Analytics & Insights feature page"""
    return render_template("analytics-feature.html")

@app.route("/visualizations")
@login_required
def visualizations_feature():
    """Visualizations feature page"""
    return render_template("visualizations-feature.html")

@app.route("/collaboration")
@login_required
def collaboration_feature():
    """Collaboration & Sharing feature page"""
    return render_template("collaboration-feature.html")

@app.route("/export")
@login_required
def export_manager():
    """Data Export Manager - Export readings to CSV, Excel, PDF"""
    return render_template("export-manager.html")

@app.route("/organizations")
@login_required
def organizations():
    """Multi-Tenant Management - Create and manage organizations"""
    return render_template("tenant-management.html")

@app.route("/team-collaboration")
@login_required
def team_collaboration():
    """Team Collaboration - Share dashboards, alerts, and comments"""
    return render_template("collaboration.html")

@app.route("/admin/performance")
@login_required
def performance_monitoring():
    """Performance Monitoring - Real-time metrics and optimization"""
    return render_template("performance-monitoring.html")

@app.route("/performance")
@login_required
def performance_feature():
    """Performance & Optimization feature page"""
    return render_template("performance-feature.html")

@app.route("/health")
@login_required
def health_feature():
    """Health Recommendations feature page"""
    return render_template("health-feature.html")

@app.route("/admin-tools")
@login_required
def admin_tools():
    """Advanced admin tools - Audit logs, sessions, retention, backups"""
    user = get_user_by_id(session.get('user_id'))
    if not user or user['role'] != 'admin':
        return redirect(url_for('dashboard'))
    return render_template("admin-tools.html")

@app.route("/debug-session")
def debug_session():
    """Debug session and auth info"""
    user_data = None
    if session.get('user_id'):
        user = get_user_by_id(session.get('user_id'))
        if user:
            user_data = {
                "id": user['id'],
                "username": user['username'],
                "role": user['role'],
            }
    
    return jsonify({
        "session_user_id": session.get('user_id'),
        "session_username": session.get('username'),
        "user_in_session": 'user_id' in session,
        "is_admin_result": is_admin(session.get('user_id')) if 'user_id' in session else None,
        "user_data": user_data
    })

@app.route("/admin")
@admin_required
def admin_dashboard():
    """Admin dashboard with statistics and user management"""
    stats = get_admin_stats()
    users = get_all_users()
    audit_logs = get_audit_logs(limit=20)
    db_info = get_database_info()
    
    return render_template("admin.html", 
                         stats=stats, 
                         users=users, 
                         audit_logs=audit_logs,
                         db_info=db_info)

@app.route("/admin/user/<int:user_id>/role/<role>", methods=["POST"])
@admin_required
def update_user_role(user_id, role):
    """Update user role (admin or user)"""
    if role not in ['user', 'admin']:
        return jsonify({'error': 'Invalid role'}), 400
    
    # Prevent self-demotion
    if user_id == session.get('user_id') and role == 'user':
        return jsonify({'error': 'Cannot remove your own admin privileges'}), 400
    
    if set_user_role(user_id, role):
        # Log the action
        admin_id = session.get('user_id')
        ip_address = request.remote_addr
        log_audit(admin_id, 'user_role_updated', 'user', user_id, 
                 'role_changed', f"user → {role}", ip_address)
        
        return jsonify({'success': True, 'role': role})
    return jsonify({'error': 'Failed to update role'}), 500

@app.route("/api/permissions", methods=["GET"])
@login_required
def get_my_permissions():
    """Get current user's permissions"""
    user_id = session.get('user_id')
    permissions = get_user_permissions(user_id)
    return jsonify({"permissions": permissions})

@app.route("/api/permissions/<int:user_id>", methods=["GET"])
@admin_required
def get_user_perms(user_id):
    """Get a specific user's permissions (admin only)"""
    permissions = get_user_permissions(user_id)
    return jsonify({"user_id": user_id, "permissions": permissions})

@app.route("/api/permissions/<int:user_id>/<permission>", methods=["POST"])
@admin_required
def add_permission(user_id, permission):
    """Grant a permission to a user"""
    valid_perms = ['view_reports', 'manage_exports', 'manage_sensors', 'manage_alerts', 'manage_users']
    
    if permission not in valid_perms:
        return jsonify({"error": f"Invalid permission. Valid: {', '.join(valid_perms)}"}), 400
    
    grant_permission(user_id, permission)
    log_audit(session.get('user_id'), 'permission_granted', 'user', user_id, 
             'permission', permission, request.remote_addr)
    
    return jsonify({"status": "ok", "permission": permission})

@app.route("/api/permissions/<int:user_id>/<permission>", methods=["DELETE"])
@admin_required
def remove_permission(user_id, permission):
    """Revoke a permission from a user"""
    revoke_permission(user_id, permission)
    log_audit(session.get('user_id'), 'permission_revoked', 'user', user_id,
             'permission', permission, request.remote_addr)
    
    return jsonify({"status": "ok", "permission": permission})

@app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_user_admin(user_id):
    """Delete a user account (admin only)"""
    # Prevent self-deletion
    if user_id == session.get('user_id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    
    admin_id = session.get('user_id')
    ip_address = request.remote_addr
    
    if delete_user_with_audit(user_id, admin_id, ip_address):
        return jsonify({'success': True})
    
    return jsonify({'error': 'Failed to delete user'}), 500

@app.route("/admin/maintenance", methods=["POST"])
@admin_required
def admin_maintenance():
    """Execute maintenance tasks"""
    task = request.json.get('task')
    admin_id = session.get('user_id')
    ip_address = request.remote_addr
    
    results = {}
    
    if task == 'cleanup_old_data':
        days = request.json.get('days', 90)
        deleted = cleanup_old_data(days)
        results['deleted_readings'] = deleted
        log_audit(admin_id, 'maintenance_cleanup_data', None, None, 
                 f'days={days}', f'deleted={deleted}', ip_address)
    
    elif task == 'cleanup_old_logs':
        days = request.json.get('days', 180)
        deleted = cleanup_old_audit_logs(days)
        results['deleted_audit_logs'] = deleted
        log_audit(admin_id, 'maintenance_cleanup_logs', None, None, 
                 f'days={days}', f'deleted={deleted}', ip_address)
    
    elif task == 'cleanup_login_history':
        days = request.json.get('days', 90)
        deleted = cleanup_old_login_history(days)
        results['deleted_logins'] = deleted
        log_audit(admin_id, 'maintenance_cleanup_logins', None, None, 
                 f'days={days}', f'deleted={deleted}', ip_address)
    
    else:
        return jsonify({'error': 'Unknown maintenance task'}), 400
    
    return jsonify({'success': True, **results})

# ================================================================================
#                        ONBOARDING ROUTES
# ================================================================================

@app.route("/onboarding")
@login_required
def onboarding_page():
    """Onboarding/tutorial page"""
    user_id = session.get('user_id')
    onboarding = get_onboarding_status(user_id)
    
    # Initialize if doesn't exist
    if not onboarding:
        init_onboarding(user_id)
        onboarding = get_onboarding_status(user_id)
    
    return render_template("onboarding.html", onboarding=onboarding)

@app.route("/api/onboarding/step/<int:step>", methods=["POST"])
@login_required
def update_onboarding_progress(step):
    """Update onboarding step"""
    user_id = session.get('user_id')
    update_onboarding_step(user_id, step)
    return jsonify({'success': True, 'step': step})

@app.route("/api/onboarding/feature/<feature>", methods=["POST"])
@login_required
def mark_feature_seen(feature):
    """Mark a feature as seen"""
    user_id = session.get('user_id')
    mark_feature_as_seen(user_id, feature)
    return jsonify({'success': True})

@app.route("/api/onboarding/complete", methods=["POST"])
@login_required
def finish_onboarding():
    """Complete onboarding"""
    user_id = session.get('user_id')
    complete_onboarding(user_id)
    return jsonify({'success': True})

@app.route("/api/onboarding/tour", methods=["POST"])
@login_required
def handle_tour():
    """Handle tour start/complete"""
    user_id = session.get('user_id')
    action = request.json.get('action')
    
    if action == 'start':
        start_tour(user_id)
    elif action == 'complete':
        complete_tour(user_id)
    
    return jsonify({'success': True})

# ================================================================================
#                        DATA EXPORT ROUTES
# ================================================================================

@app.route("/api/export/json")
@login_required
@limiter.limit("10 per minute")
def export_json():
    """Export CO₂ data as JSON"""
    user_id = session.get('user_id')
    days = request.args.get('days', 30, type=int)

    db_source = resolve_source_param(allow_sim=False, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)
    
    db = get_db()
    readings = db.execute(f"""
        SELECT timestamp, ppm FROM co2_readings
        WHERE timestamp >= datetime('now', '-' || ? || ' days')
        AND {source_clause}
        ORDER BY timestamp DESC
    """, (days, *source_params)).fetchall()
    db.close()
    
    data = {
        'export_date': datetime.now(UTC).isoformat(),
        'days': days,
        'count': len(readings),
        'readings': [{'timestamp': r['timestamp'], 'ppm': r['ppm']} for r in readings]
    }
    
    return jsonify(data)

@app.route("/api/export/csv")
@login_required
@limiter.limit("10 per minute")
def export_csv():
    """Export CO₂ data as CSV"""
    user_id = session.get('user_id')
    days = request.args.get('days', 30, type=int)

    db_source = resolve_source_param(allow_sim=False, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)
    
    db = get_db()
    readings = db.execute(f"""
        SELECT timestamp, ppm FROM co2_readings
        WHERE timestamp >= datetime('now', '-' || ? || ' days')
        AND {source_clause}
        ORDER BY timestamp DESC
    """, (days, *source_params)).fetchall()
    db.close()
    
    # Generate CSV
    csv_content = "timestamp,ppm\n"
    for row in readings:
        csv_content += f"{row['timestamp']},{row['ppm']}\n"
    
    response = make_response(csv_content)
    response.headers['Content-Disposition'] = f'attachment; filename="co2_export_{days}d.csv"'
    response.headers['Content-Type'] = 'text/csv'
    
    return response

@app.route("/api/export/excel")
@login_required
@limiter.limit("10 per minute")
def export_excel():
    """Export CO₂ data as Excel"""
    user_id = session.get('user_id')
    days = request.args.get('days', 30, type=int)

    db_source = resolve_source_param(allow_sim=False, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)
    
    db = get_db()
    readings = db.execute(f"""
        SELECT timestamp, ppm FROM co2_readings
        WHERE timestamp >= datetime('now', '-' || ? || ' days')
        AND {source_clause}
        ORDER BY timestamp DESC
    """, (days, *source_params)).fetchall()
    db.close()
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            return jsonify({'error': 'Failed to create Excel workbook'}), 500
        
        ws.title = "CO₂ Data"
        
        # Headers
        ws['A1'] = "Timestamp"
        ws['B1'] = "PPM"
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="3DD98F", end_color="3DD98F", fill_type="solid")
        
        # Data
        for idx, row in enumerate(readings, start=2):
            ws[f'A{idx}'] = row['timestamp']
            ws[f'B{idx}'] = row['ppm']
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename="co2_export_{days}d.xlsx"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    except ImportError:
        return jsonify({'error': 'Excel support not installed. Install openpyxl: pip install openpyxl'}), 400

@app.route("/api/export/schedule", methods=["POST"])
@login_required
def schedule_export():
    """Setup scheduled exports"""
    user_id = session.get('user_id')
    format = request.json.get('format', 'csv')
    frequency = request.json.get('frequency', 'weekly')
    
    if format not in ['csv', 'json', 'excel']:
        return jsonify({'error': 'Invalid format'}), 400
    
    if frequency not in ['daily', 'weekly', 'monthly']:
        return jsonify({'error': 'Invalid frequency'}), 400
    
    create_scheduled_export(user_id, format, frequency)
    return jsonify({'success': True, 'format': format, 'frequency': frequency})

@app.route("/api/export/scheduled")
@login_required
def get_scheduled_exports():
    """Get user's scheduled exports"""
    user_id = session.get('user_id')
    exports = get_user_scheduled_exports(user_id)
    
    return jsonify({
        'exports': [dict(e) for e in exports]
    })

@app.route("/api/export/scheduled/<int:export_id>", methods=["DELETE"])
@login_required
def remove_scheduled_export(export_id):
    """Remove a scheduled export"""
    user_id = session.get('user_id')
    
    # Verify ownership
    db = get_db()
    export = db.execute(
        "SELECT user_id FROM scheduled_exports WHERE id = ?",
        (export_id,)
    ).fetchone()
    db.close()
    
    if not export or export['user_id'] != user_id:
        return jsonify({'error': 'Not found'}), 404
    
    delete_scheduled_export(export_id)
    return jsonify({'success': True})

@app.route("/api/import/csv", methods=["POST"])
@login_required
@limiter.limit("5 per minute")
def import_csv():
    """Import CO₂ readings from CSV file"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if not file or not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    
    filename = secure_filename(file.filename)
    if not filename or not filename.lower().endswith('.csv'):
        return jsonify({'error': 'File must be CSV format'}), 400
    
    try:
        # Parse CSV
        stream = io.TextIOWrapper(file.stream, encoding='utf-8')
        reader = csv.DictReader(stream)
        readings = []
        
        for row in reader:
            if row.get('timestamp') and row.get('ppm'):
                readings.append({
                    'timestamp': row['timestamp'],
                    'ppm': row['ppm']
                })
        
        if not readings:
            return jsonify({'error': 'No valid readings found in CSV'}), 400
        
        # Import
        result = import_csv_readings(readings)
        
        # Log the import
        log_audit(session.get('user_id'), 'CSV_IMPORT', 'data', 0,
                 'imported_count', str(result['imported']), request.remote_addr)
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': f'Failed to parse CSV: {str(e)}'}), 400

@app.route("/api/import/stats")
@admin_required
def import_stats():
    """Get data import statistics"""
    stats = get_csv_import_stats()
    return jsonify(stats)

# ================================================================================
#                        SENSOR MANAGEMENT ROUTES
# ================================================================================

_sensor_mode = os.getenv('USE_SCD30', '1')  # Default to real if available
_sensor_last_read = 0

@app.route("/api/sensor/status")
@login_required
def sensor_status():
    """Get current sensor mode and availability status"""
    global _sensor_mode, _sensor_last_read
    
    try:
        # Try to import and check if SCD30 is available
        from app.sensors.scd30 import SCD30
        scd30 = SCD30()
        
        # Try to read to verify it's working
        try:
            reading = scd30.read()
            available = reading is not None
            _sensor_last_read = time.time()
        except Exception:
            available = False
    except Exception:
        available = False
    
    mode = "real" if _sensor_mode != "0" and available else "simulation"
    
    return jsonify({
        'mode': mode,
        'available': available,
        'last_read': _sensor_last_read,
        'driver': 'SCD30' if available else 'None'
    })

@app.route("/api/sensor/mode", methods=["POST"])
@login_required
def set_sensor_mode():
    """Set sensor mode (real or simulation)"""
    global _sensor_mode
    
    data = request.get_json()
    mode = data.get('mode', 'simulation')
    
    if mode not in ['real', 'simulation']:
        return jsonify({'error': 'Invalid mode. Must be "real" or "simulation"'}), 400
    
    # Update environment for future reads
    _sensor_mode = "1" if mode == "real" else "0"
    os.environ['USE_SCD30'] = _sensor_mode
    
    # Log the change
    try:
        log_audit(session.get('user_id'), 'SENSOR_MODE_CHANGE', 'sensor',
                 0, 'new_mode', mode, request.remote_addr)
    except Exception:
        pass
    
    return jsonify({
        'success': True,
        'mode': mode,
        'message': f'Mode capteur défini à: {mode}'
    })

@app.route("/api/sensor/test", methods=["POST"])
@login_required
def test_sensor():
    """Test SCD30 sensor connection"""
    data = request.get_json()
    bus = int(data.get('bus', 1))
    address = data.get('address', '0x61')
    
    # Convert hex address to int if needed
    if isinstance(address, str):
        address = int(address, 16)
    
    try:
        from app.sensors.scd30 import SCD30
        scd30 = SCD30(bus=bus, address=address)
        reading = scd30.read()
        
        if reading and 'co2' in reading:
            return jsonify({
                'success': True,
                'co2': round(reading['co2'], 2),
                'temperature': round(reading.get('temperature', 0), 2),
                'humidity': round(reading.get('humidity', 0), 2),
                'message': 'Capteur détecté avec succès'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Capteur détecté mais pas de lecture disponible'
            }), 400
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Impossible de se connecter au capteur: {str(e)}'
        }), 400

# ================================================================================
#                        MULTI-SENSOR API ENDPOINTS
# ================================================================================

@app.route("/api/sensors", methods=["GET"])
@login_required
def get_sensors_list():
    """Get all sensors for current user"""
    try:
        user_id = session.get('user_id')
        sensors = get_user_sensors(user_id)
        return jsonify(sensors)
    except Exception as e:
        print(f"Error getting sensors: {e}")
        return jsonify({'error': str(e)}), 500

@app.route("/api/sensors", methods=["POST"])
@login_required
def create_new_sensor():
    """Create a new sensor"""
    try:
        user_id = session.get('user_id')
        data = request.get_json()
        
        name = data.get('name', '').strip()
        sensor_type = data.get('type', 'scd30')
        interface = data.get('interface', 'i2c')
        config = data.get('config', {})
        
        if not name:
            return jsonify({'error': 'Sensor name is required'}), 400
        
        sensor_id = create_sensor(user_id, name, sensor_type, interface, config)
        
        if not sensor_id:
            return jsonify({'error': 'Sensor name already exists for this user'}), 400
        
        # Log action
        try:
            log_audit(user_id, 'SENSOR_CREATED', 'sensor', sensor_id, None, f'{name} ({sensor_type})', request.remote_addr)
        except:
            pass
        
        sensor = get_sensor_by_id(sensor_id, user_id)
        return jsonify(sensor), 201
    except Exception as e:
        print(f"Error creating sensor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route("/api/sensors/<int:sensor_id>", methods=["PUT"])
@login_required
def update_sensor_config(sensor_id):
    """Update sensor configuration"""
    try:
        user_id = session.get('user_id')
        data = request.get_json()
        
        # Verify ownership
        sensor = get_sensor_by_id(sensor_id, user_id)
        if not sensor:
            return jsonify({'error': 'Sensor not found or not accessible'}), 404
        
        # Update fields if provided
        name = data.get('name')
        sensor_type = data.get('type')
        interface = data.get('interface')
        config = data.get('config')
        active = data.get('active')
        
        update_sensor(sensor_id, user_id, 
                     name=name,
                     sensor_type=sensor_type,
                     interface=interface,
                     config=config,
                     active=active)
        
        # Log action
        try:
            log_audit(user_id, 'SENSOR_UPDATED', 'sensor', sensor_id, None, name or sensor['name'], request.remote_addr)
        except:
            pass
        
        updated_sensor = get_sensor_by_id(sensor_id, user_id)
        return jsonify(updated_sensor)
    except Exception as e:
        print(f"Error updating sensor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route("/api/sensors/<int:sensor_id>", methods=["DELETE"])
@login_required
def delete_sensor_endpoint(sensor_id):
    """Delete a sensor"""
    try:
        user_id = session.get('user_id')
        
        # Verify ownership
        sensor = get_sensor_by_id(sensor_id, user_id)
        if not sensor:
            return jsonify({'error': 'Sensor not found or not accessible'}), 404
        
        delete_sensor(sensor_id, user_id)
        
        # Log action
        try:
            log_audit(user_id, 'SENSOR_DELETED', 'sensor', sensor_id, sensor['name'], None, request.remote_addr)
        except:
            pass
        
        return jsonify({'success': True, 'message': 'Sensor deleted'})
    except Exception as e:
        print(f"Error deleting sensor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route("/api/sensors/test", methods=["POST"])
@login_required
def test_sensor_connection():
    """Test sensor connection with provided configuration"""
    try:
        data = request.get_json()
        sensor_type = data.get('type', 'scd30')
        interface = data.get('interface', 'i2c')
        config = data.get('config', {})
        
        # Test SCD30 sensor
        if sensor_type == 'scd30' and interface == 'i2c':
            try:
                from app.sensors.scd30 import SCD30
                bus = config.get('bus', 1)
                address = config.get('address', '0x61')
                
                # Convert hex address if needed
                if isinstance(address, str):
                    address = int(address, 16)
                
                scd30 = SCD30(bus=int(bus), address=address)
                reading = scd30.read()
                
                if reading and 'co2' in reading:
                    return jsonify({
                        'success': True,
                        'co2': round(reading.get('co2', 0), 2),
                        'temperature': round(reading.get('temperature', 0), 2),
                        'humidity': round(reading.get('humidity', 0), 2),
                        'message': 'Capteur détecté'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'error': 'Capteur trouvé mais pas de lecture'
                    })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': f'Erreur SCD30: {str(e)}'
                })
        
        # Add support for other sensor types here
        elif sensor_type == 'mhz19':
            return jsonify({
                'success': False,
                'error': 'MH-Z19 support en développement'
            })
        
        else:
            return jsonify({
                'success': False,
                'error': f'Type de capteur non supporté: {sensor_type}'
            })
            
    except Exception as e:
        print(f"Error testing sensor: {e}")
        return jsonify({
            'success': False,
            'error': f'Erreur de test: {str(e)}'
        }), 500

# ================================================================================
#                    SENSOR READINGS & THRESHOLDS
# ================================================================================

@app.route("/api/sensor/<int:sensor_id>/readings")
@login_required
def get_sensor_readings_endpoint(sensor_id):
    """Get readings for a specific sensor (last 24 hours)"""
    try:
        user_id = session.get('user_id')
        hours = request.args.get('hours', 24, type=int)
        
        # Verify ownership
        sensor = get_sensor_by_id(sensor_id, user_id)
        if not sensor:
            return jsonify({'error': 'Sensor not found'}), 404
        
        readings = get_sensor_readings(sensor_id, hours)
        latest = get_sensor_latest_reading(sensor_id)
        
        return jsonify({
            'sensor_id': sensor_id,
            'sensor_name': sensor['name'],
            'readings': readings,
            'latest': latest,
            'count': len(readings)
        })
    except Exception as e:
        print(f"Error getting readings: {e}")
        return jsonify({'error': str(e)}), 500

@app.route("/api/sensor/<int:sensor_id>/thresholds", methods=["GET"])
@login_required
def get_sensor_thresholds_endpoint(sensor_id):
    """Get thresholds for a specific sensor"""
    try:
        user_id = session.get('user_id')
        
        # Verify ownership
        sensor = get_sensor_by_id(sensor_id, user_id)
        if not sensor:
            return jsonify({'error': 'Sensor not found'}), 404
        
        thresholds = {
            'good': sensor.get('good_threshold', 800),
            'warning': sensor.get('warning_threshold', 1000),
            'critical': sensor.get('critical_threshold', 1200)
        }
        
        return jsonify(thresholds)
    except Exception as e:
        print(f"Error getting thresholds: {e}")
        return jsonify({'error': str(e)}), 500

@app.route("/api/sensor/<int:sensor_id>/thresholds", methods=["PUT"])
@login_required
def update_sensor_thresholds_endpoint(sensor_id):
    """Update thresholds for a specific sensor"""
    try:
        user_id = session.get('user_id')
        data = request.get_json()
        
        # Verify ownership
        sensor = get_sensor_by_id(sensor_id, user_id)
        if not sensor:
            return jsonify({'error': 'Sensor not found'}), 404
        
        good = data.get('good')
        warning = data.get('warning')
        critical = data.get('critical')
        
        success = update_sensor_thresholds(
            sensor_id, user_id,
            good=good, warning=warning, critical=critical
        )
        
        if success:
            # Log the change
            try:
                log_audit(user_id, 'SENSOR_THRESHOLDS_UPDATED', 'sensor', sensor_id,
                         None, f'{sensor["name"]}: {good}/{warning}/{critical}', request.remote_addr)
            except:
                pass
            
            return jsonify({'success': True, 'message': 'Thresholds updated'})
        else:
            return jsonify({'error': 'Failed to update thresholds'}), 400
    except Exception as e:
        print(f"Error updating thresholds: {e}")
        return jsonify({'error': str(e)}), 500

# ================================================================================
#                        ANALYTICS ROUTES (ENHANCED)
# ================================================================================

@app.route("/api/analytics/weekcompare")
@login_required
def analytics_week_compare():
    """Get week-over-week comparison data"""
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)

    db = get_db()
    
    # Current week
    current_week = db.execute(f"""
        SELECT DATE(timestamp) as date, AVG(ppm) as avg_ppm, MAX(ppm) as max_ppm, MIN(ppm) as min_ppm, COUNT(*) as count
        FROM co2_readings
        WHERE timestamp >= datetime('now', '-7 days')
        AND {source_clause}
        GROUP BY DATE(timestamp)
        ORDER BY date
    """, source_params).fetchall()
    
    # Previous week
    prev_week = db.execute(f"""
        SELECT DATE(timestamp) as date, AVG(ppm) as avg_ppm, MAX(ppm) as max_ppm, MIN(ppm) as min_ppm, COUNT(*) as count
        FROM co2_readings
        WHERE timestamp >= datetime('now', '-14 days') AND timestamp < datetime('now', '-7 days')
        AND {source_clause}
        GROUP BY DATE(timestamp)
        ORDER BY date
    """, source_params).fetchall()
    
    db.close()
    
    return jsonify({
        'current_week': [dict(row) for row in current_week],
        'previous_week': [dict(row) for row in prev_week]
    })

@app.route("/api/analytics/trend")
@login_required
def analytics_trend():
    """Get trend analysis (rising/stable/falling)"""
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)
    
    db = get_db()
    
    # For imported data, show ALL data; for live/sim, show last 7 days
    if db_source == 'import':
        data = db.execute(f"""
            SELECT 
                strftime('%Y-%m-%d %H:00', timestamp) as hour,
                AVG(ppm) as avg_ppm,
                COUNT(*) as readings
            FROM co2_readings
            WHERE {source_clause}
            GROUP BY hour
            ORDER BY hour DESC
            LIMIT 168
        """, source_params).fetchall()
    else:
        data = db.execute(f"""
            SELECT 
                strftime('%Y-%m-%d %H:00', timestamp) as hour,
                AVG(ppm) as avg_ppm,
                COUNT(*) as readings
            FROM co2_readings
            WHERE timestamp >= datetime('now', '-7 days')
            AND {source_clause}
            GROUP BY hour
            ORDER BY hour DESC
            LIMIT 168
        """, source_params).fetchall()
    
    db.close()
    
    if len(data) < 2:
        return jsonify({'trend': 'insufficient_data', 'data': []})
    
    # Calculate trend
    recent_avg = sum(d['avg_ppm'] for d in data[:24]) / min(24, len(data))
    
    # Check if we have enough older data
    older_data = data[24:48]
    if len(older_data) == 0:
        # Not enough data for comparison
        return jsonify({
            'trend': 'insufficient_data',
            'recent_avg': round(recent_avg, 1),
            'older_avg': 0,
            'percent_change': 0,
            'data': [dict(d) for d in data]
        })
    
    older_avg = sum(d['avg_ppm'] for d in older_data) / len(older_data)
    
    if older_avg == 0:
        trend = 'insufficient_data'
    else:
        percent_change = ((recent_avg - older_avg) / older_avg) * 100
        if percent_change > 5:
            trend = 'rising'
        elif percent_change < -5:
            trend = 'falling'
        else:
            trend = 'stable'
    
    return jsonify({
        'trend': trend,
        'recent_avg': round(recent_avg, 1),
        'older_avg': round(older_avg, 1),
        'percent_change': round(((recent_avg - older_avg) / older_avg) * 100, 1) if older_avg else 0,
        'data': [dict(d) for d in data]
    })

@app.route("/api/analytics/custom")
@login_required
def analytics_custom_range():
    """Get data for custom date range"""
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    
    if not start_date or not end_date:
        return jsonify({'error': 'start and end dates required'}), 400
    
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)

    db = get_db()
    
    readings = db.execute(f"""
        SELECT timestamp, ppm FROM co2_readings
        WHERE DATE(timestamp) >= ? AND DATE(timestamp) <= ?
        AND {source_clause}
        ORDER BY timestamp
    """, (start_date, end_date, *source_params)).fetchall()
    
    # Calculate stats
    stats = db.execute(f"""
        SELECT 
            COUNT(*) as count,
            AVG(ppm) as avg,
            MIN(ppm) as min,
            MAX(ppm) as max
        FROM co2_readings
        WHERE DATE(timestamp) >= ? AND DATE(timestamp) <= ?
        AND {source_clause}
    """, (start_date, end_date, *source_params)).fetchone()
    
    db.close()
    
    return jsonify({
        'readings': [dict(r) for r in readings],
        'stats': dict(stats) if stats else {'count': 0, 'avg': 0, 'min': 0, 'max': 0}
    })

@app.route("/api/analytics/compare-periods")
@login_required
def compare_periods():
    """Compare CO₂ data between two time periods (e.g., this week vs last week)"""
    period_type = request.args.get('type', 'week')  # week, month, or custom
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)
    
    db = get_db()
    
    if period_type == 'week':
        # For imported data, compare first half vs second half of dataset
        if db_source == 'import':
            # Get date range of imported data
            date_range = db.execute(f"""
                SELECT MIN(timestamp) as min_date, MAX(timestamp) as max_date
                FROM co2_readings WHERE {source_clause}
            """, source_params).fetchone()
            
            if date_range and date_range['min_date'] and date_range['max_date']:
                # Split in half
                current_data = db.execute(f"""
                    SELECT AVG(ppm) as avg, MIN(ppm) as min, MAX(ppm) as max, COUNT(*) as count
                    FROM co2_readings
                    WHERE {source_clause} AND timestamp >= (SELECT datetime((julianday(MIN(timestamp)) + julianday(MAX(timestamp))) / 2) FROM co2_readings WHERE {source_clause})
                """, (*source_params, *source_params)).fetchone()
                
                previous_data = db.execute(f"""
                    SELECT AVG(ppm) as avg, MIN(ppm) as min, MAX(ppm) as max, COUNT(*) as count
                    FROM co2_readings
                    WHERE {source_clause} AND timestamp < (SELECT datetime((julianday(MIN(timestamp)) + julianday(MAX(timestamp))) / 2) FROM co2_readings WHERE {source_clause})
                """, (*source_params, *source_params)).fetchone()
            else:
                current_data = {'avg': 0, 'min': 0, 'max': 0, 'count': 0}
                previous_data = {'avg': 0, 'min': 0, 'max': 0, 'count': 0}
        else:
            # Compare this week to last week
            current_data = db.execute(f"""
                SELECT AVG(ppm) as avg, MIN(ppm) as min, MAX(ppm) as max, COUNT(*) as count
                FROM co2_readings
                WHERE timestamp >= datetime('now', '-7 days')
                AND {source_clause}
            """, source_params).fetchone()
            
            previous_data = db.execute(f"""
                SELECT AVG(ppm) as avg, MIN(ppm) as min, MAX(ppm) as max, COUNT(*) as count
                FROM co2_readings
                WHERE timestamp >= datetime('now', '-14 days') AND timestamp < datetime('now', '-7 days')
                AND {source_clause}
            """, source_params).fetchone()
        
    elif period_type == 'month':
        # Compare this month to last month
        current_data = db.execute(f"""
            SELECT AVG(ppm) as avg, MIN(ppm) as min, MAX(ppm) as max, COUNT(*) as count
            FROM co2_readings
            WHERE timestamp >= datetime('now', 'start of month')
            AND {source_clause}
        """, source_params).fetchone()
        
        previous_data = db.execute(f"""
            SELECT AVG(ppm) as avg, MIN(ppm) as min, MAX(ppm) as max, COUNT(*) as count
            FROM co2_readings
            WHERE timestamp >= datetime('now', '-1 month', 'start of month')
            AND timestamp < datetime('now', 'start of month')
            AND {source_clause}
        """, source_params).fetchone()
    else:
        return jsonify({'error': 'Invalid period_type'}), 400
    
    db.close()
    
    # Calculate differences
    def calc_diff(current, previous, field):
        if previous[field] is None or previous[field] == 0:
            return 0
        return round(((current[field] - previous[field]) / previous[field]) * 100, 1)
    
    return jsonify({
        'period': period_type,
        'current': dict(current_data) if current_data else {},
        'previous': dict(previous_data) if previous_data else {},
        'differences': {
            'avg_percent': calc_diff(current_data, previous_data, 'avg'),
            'min_percent': calc_diff(current_data, previous_data, 'min'),
            'max_percent': calc_diff(current_data, previous_data, 'max'),
        }
    })

@app.route("/api/analytics/daily-comparison")
@login_required
def daily_comparison():
    """Get daily averages for the last 30 days for trend visualization"""
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)
    
    db = get_db()
    
    # For imported data, show ALL data regardless of date range
    # For live/sim data, show last 30 days
    if db_source == 'import':
        days_data = db.execute(f"""
            SELECT 
                DATE(timestamp) as date,
                AVG(ppm) as avg_ppm,
                MIN(ppm) as min_ppm,
                MAX(ppm) as max_ppm,
                COUNT(*) as readings
            FROM co2_readings
            WHERE {source_clause}
            GROUP BY DATE(timestamp)
            ORDER BY date ASC
        """, source_params).fetchall()
    else:
        days_data = db.execute(f"""
            SELECT 
                DATE(timestamp) as date,
                AVG(ppm) as avg_ppm,
                MIN(ppm) as min_ppm,
                MAX(ppm) as max_ppm,
                COUNT(*) as readings
            FROM co2_readings
            WHERE timestamp >= datetime('now', '-30 days')
            AND {source_clause}
            GROUP BY DATE(timestamp)
            ORDER BY date ASC
        """, source_params).fetchall()
    
    db.close()
    
    return jsonify({
        'days': 30,
        'data': [dict(d) for d in days_data]
    })

@app.route("/api/analytics/report/pdf")
@login_required
def generate_pdf_report():
    """Generate PDF report with charts"""
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    
    if not start_date or not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    
    db_source = resolve_source_param(allow_sim=False, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)

    db = get_db()
    
    # Get data
    readings = db.execute(f"""
        SELECT DATE(timestamp) as date, AVG(ppm) as avg_ppm, MAX(ppm) as max_ppm, MIN(ppm) as min_ppm
        FROM co2_readings
        WHERE DATE(timestamp) >= ? AND DATE(timestamp) <= ?
        AND {source_clause}
        GROUP BY DATE(timestamp)
        ORDER BY date
    """, (start_date, end_date, *source_params)).fetchall()
    
    stats = db.execute(f"""
        SELECT 
            COUNT(*) as count,
            AVG(ppm) as avg,
            MIN(ppm) as min,
            MAX(ppm) as max
        FROM co2_readings
        WHERE DATE(timestamp) >= ? AND DATE(timestamp) <= ?
        AND {source_clause}
    """, (start_date, end_date, *source_params)).fetchone()
    
    db.close()
    
    # Build HTML for PDF
    html_content = f"""
    <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #3dd98f; }}
                .stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; margin: 20px 0; }}
                .stat {{ padding: 15px; background: #f0f0f0; border-radius: 8px; text-align: center; }}
                .stat-value {{ font-size: 24px; font-weight: bold; color: #3dd98f; }}
                .stat-label {{ color: #666; margin-top: 5px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #3dd98f; color: white; }}
            </style>
        </head>
        <body>
            <h1>🌬️ CO₂ Monitoring Report</h1>
            <p>Period: {start_date} to {end_date}</p>
            
            <div class="stats">
                <div class="stat">
                    <div class="stat-value">{stats['count']}</div>
                    <div class="stat-label">Readings</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{round(stats['avg'], 1)}</div>
                    <div class="stat-label">Avg PPM</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{round(stats['min'], 0)}</div>
                    <div class="stat-label">Min PPM</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{round(stats['max'], 0)}</div>
                    <div class="stat-label">Max PPM</div>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Average PPM</th>
                        <th>Min PPM</th>
                        <th>Max PPM</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for row in readings:
        html_content += f"""
                    <tr>
                        <td>{row['date']}</td>
                        <td>{round(row['avg_ppm'], 1)}</td>
                        <td>{round(row['min_ppm'], 0)}</td>
                        <td>{round(row['max_ppm'], 0)}</td>
                    </tr>
        """
    
    html_content += """
                </tbody>
            </table>
        </body>
    </html>
    """
    
    # Generate PDF
    if not WEASYPRINT_AVAILABLE or HTML is None:
        return jsonify({
            'error': 'WeasyPrint is not available. Install system dependencies: https://doc.courtbouillon.org/weasyprint/stable/first_steps.html#installation'
        }), 500
    try:
        pdf = HTML(string=html_content).write_pdf()
        response = make_response(pdf)
        response.headers['Content-Disposition'] = f'attachment; filename="report_{start_date}_to_{end_date}.pdf"'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF: {str(e)}'}), 500

@app.route("/api/history/<range>")
def history_range(range):
    db = get_db()

    if range == "today":
        rows = db.execute("""
            SELECT ppm, timestamp
            FROM co2_readings
            WHERE date(timestamp) = date('now')
            ORDER BY timestamp
        """).fetchall()

    elif range == "7d":
        rows = db.execute("""
            SELECT ppm, timestamp
            FROM co2_readings
            WHERE timestamp >= datetime('now', '-7 days')
            ORDER BY timestamp
        """).fetchall()

    elif range == "30d":
        rows = db.execute("""
            SELECT ppm, timestamp
            FROM co2_readings
            WHERE timestamp >= datetime('now', '-30 days')
            ORDER BY timestamp
        """).fetchall()

    else:
        db.close()
        return jsonify({"error": "Invalid range"}), 400

    db.close()
    return jsonify([dict(r) for r in rows])

# 3. API ROUTES
def get_latest_real_reading(max_age_minutes: int = 1):
    """Return the most recent hardware reading if it is fresh enough (default 1 min)."""
    db = get_db()
    row = db.execute(
        """
        SELECT ppm, temperature, humidity, timestamp, source
        FROM co2_readings
        WHERE source IN ('sensor','live_real')
        ORDER BY timestamp DESC
        LIMIT 1
        """
    ).fetchone()
    db.close()

    if not row:
        return None

    ts_raw = row["timestamp"]
    if ts_raw:
        try:
            ts_dt = datetime.fromisoformat(ts_raw.replace("Z", "+00:00"))
            if ts_dt < datetime.now(UTC) - timedelta(minutes=max_age_minutes):
                return None
        except Exception:
            # If parsing fails, treat as stale to avoid surfacing ghost values
            return None

    return dict(row)


def build_live_payload(settings=None):
    """Centralized live-reading builder (used by HTTP and WebSocket)."""
    settings = settings or load_settings()

    # Respect paused state first
    if not settings.get("analysis_running", True):
        return settings, {
            "analysis_running": False,
            "ppm": None,
            "reason": "paused",
            "timestamp": datetime.now(UTC).isoformat()
        }

    # Simulator data must NOT feed the live UI; only the simulator page uses /api/simulator/latest
    # If simulate_live is flagged true, treat it as no_sensor for the live endpoint/UI.
    if settings.get("simulate_live", False):
        return settings, {
            "analysis_running": False,
            "ppm": None,
            "reason": "no_sensor",
            "timestamp": datetime.now(UTC).isoformat()
        }

    # Real sensor path: use freshest non-simulated reading
    latest = get_latest_real_reading()
    if not latest:
        return settings, {
            "analysis_running": False,
            "ppm": None,
            "reason": "no_sensor",
            "timestamp": datetime.now(UTC).isoformat()
        }

    return settings, {
        "analysis_running": True,
        "ppm": latest.get("ppm"),
        "temp": latest.get("temperature"),
        "humidity": latest.get("humidity"),
        "timestamp": latest.get("timestamp", datetime.now(UTC).isoformat())
    }


@app.route("/api/live/latest")
@limiter.exempt
def api_live_latest():
    _, payload = build_live_payload()
    resp = make_response(jsonify(payload))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/api/readings", methods=["POST", "GET"])
@limiter.exempt
def api_readings_ingest():
    """Ingest real sensor readings (POST) or fetch recent readings (GET)."""
    if request.method == "GET":
        days = request.args.get("days", default=1, type=int)
        db_source = resolve_source_param(allow_sim=True, allow_import=True)
        source_clause, source_params = build_source_filter(db_source)
        db = get_db()
        rows = db.execute(
            f"""
            SELECT ppm, temperature, humidity, timestamp, source
            FROM co2_readings
            WHERE timestamp >= datetime('now', ?)
            AND {source_clause}
            ORDER BY timestamp DESC
            """,
            (f"-{days} days", *source_params)
        ).fetchall()
        db.close()
        return jsonify([dict(r) for r in rows])

    data = request.get_json(silent=True) or {}
    ppm = data.get("ppm")
    if ppm is None:
        return jsonify({"error": "ppm is required"}), 400

    temp = data.get("temp", data.get("temperature"))
    humidity = data.get("humidity")

    try:
        ppm_int = int(ppm)
    except Exception:
        return jsonify({"error": "ppm must be numeric"}), 400

    save_reading(ppm_int, temp, humidity, source="sensor", persist=True)
    return jsonify({"success": True})


@app.route("/api/latest")
def api_latest():
    # Backward-compatible alias
    return api_live_latest()


@app.route("/api/simulator/latest")
@limiter.exempt
def api_simulator_latest():
    """Simulator-only feed that stays independent from the live pipeline."""
    settings = load_settings()

    data = generate_co2_data(settings.get("realistic_mode", True))
    ppm = data["co2"]
    temp = data.get("temp")
    humidity = data.get("humidity")

    # Persist simulator traffic as source="sim" so analytics can display it
    save_reading(ppm, temp, humidity, source="sim", persist=True)

    resp = make_response(jsonify({
        "analysis_running": True,
        "ppm": ppm,
        "temp": temp,
        "humidity": humidity,
        "timestamp": datetime.now(UTC).isoformat()
    }))
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.route("/api/history/today")
def api_history_today():
    source_raw = request.args.get("source", "real")

    db = get_db()
    if source_raw == "all":
        rows = db.execute(
            """
            SELECT ppm, temperature, humidity, timestamp, source
            FROM co2_readings
            WHERE date(timestamp) = date('now')
            ORDER BY timestamp
            """
        ).fetchall()
    else:
        db_source = resolve_source_param(default="live", allow_sim=True, allow_import=True)
        source_clause, source_params = build_source_filter(db_source)
        rows = db.execute(
            f"""
            SELECT ppm, temperature, humidity, timestamp, source
            FROM co2_readings
            WHERE date(timestamp) = date('now') AND {source_clause}
            ORDER BY timestamp
            """,
            source_params
        ).fetchall()

    db.close()

    return jsonify([dict(r) for r in rows])

def get_today_history(db_source="live"):
    source_clause, source_params = build_source_filter(db_source)

    db = get_db()
    rows = db.execute(f"""
        SELECT ppm, timestamp
        FROM co2_readings
        WHERE date(timestamp) = date('now') AND {source_clause}
        ORDER BY timestamp
    """, source_params).fetchall()
    db.close()

    return [dict(r) for r in rows]


@limiter.exempt
@app.route("/api/settings", methods=["GET", "POST", "DELETE"])
def api_settings():
    if request.method == "POST":
        data = request.json
        
        # Map warning_threshold to bad_threshold for backend compatibility
        if 'warning_threshold' in data and 'bad_threshold' not in data:
            data['bad_threshold'] = data['warning_threshold']
        if 'critical_threshold' in data and 'alert_threshold' not in data:
            data['alert_threshold'] = data['critical_threshold']
        
        save_settings(data)
        
        # Broadcast updated settings to all WebSocket clients
        saved_settings = load_settings()
        socketio.emit('settings_update', saved_settings)
        
        return jsonify({"status": "ok", "settings": saved_settings})

    if request.method == "DELETE":
        db = get_db()
        db.execute("DELETE FROM settings")
        db.commit()
        db.close()
        
        # Broadcast reset to all WebSocket clients
        socketio.emit('settings_update', DEFAULT_SETTINGS)
        
        return jsonify(DEFAULT_SETTINGS)

    # GET: Return settings with frontend-compatible field names
    settings = load_settings()
    
    # Map backend fields to frontend expected names
    if 'bad_threshold' in settings and 'warning_threshold' not in settings:
        settings['warning_threshold'] = settings['bad_threshold']
    if 'alert_threshold' in settings and 'critical_threshold' not in settings:
        settings['critical_threshold'] = settings['alert_threshold']
    
    return jsonify(settings)

@app.route("/api/user/profile")
@login_required
def api_user_profile():
    """Get current user's profile information"""
    user_id = session['user_id']
    user = get_user_by_id(user_id)
    
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    return jsonify({
        "id": user['id'],
        "username": user['username'],
        "email": user['email'],
        "role": "Admin" if is_admin(user_id) else "Utilisateur",
        "created_at": user.get('created_at'),
    })

@app.route("/api/user/change-password", methods=["POST"])
@login_required
def api_change_password():
    """Change user password"""
    user_id = session['user_id']
    data = request.json
    
    if not data:
        return jsonify({"error": "No JSON data provided"}), 400
    
    current_password = data.get("current_password")
    new_password = data.get("new_password")
    
    if not current_password or not new_password:
        return jsonify({"error": "Missing required fields"}), 400
    
    if len(new_password) < 8:
        return jsonify({"error": "Password must be at least 8 characters"}), 400
    
    user = get_user_by_id(user_id)
    if not user or not check_password_hash(user['password_hash'], current_password):
        return jsonify({"error": "Current password is incorrect"}), 401
    
    # Update password
    db = get_db()
    hashed = generate_password_hash(new_password)
    db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (hashed, user_id))
    db.commit()
    db.close()
    
    log_audit(user_id, "UPDATE", "Password changed")
    
    return jsonify({"status": "ok", "message": "Password changed successfully"})

@app.route("/api/admin/database-info")
@login_required
def api_database_info():
    """Get database information (admin only)"""
    user_id = session['user_id']
    
    if not is_admin(user_id):
        return jsonify({"error": "Admin access required"}), 403
    
    try:
        import os
        import sqlite3
        from pathlib import Path
        
        # Use same database path as database.py
        main_db_path = Path("../data/aerium.sqlite")
        site_db_path = Path("data/aerium.sqlite")
        
        if main_db_path.exists():
            db_path = str(main_db_path)
        else:
            db_path = str(site_db_path)
        
        # Get database file info
        if os.path.exists(db_path):
            size = os.path.getsize(db_path)
            modified = os.path.getmtime(db_path)
        else:
            return jsonify({"error": "Database file not found"}), 404
        
        # Get database schema
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cursor.fetchall()]
        
        # Build schema string
        schema_str = "Tables:\n"
        for table in tables:
            try:
                cursor.execute(f"PRAGMA table_info({table})")
                columns = cursor.fetchall()
                schema_str += f"\n{table}:\n"
                for col in columns:
                    schema_str += f"  - {col[1]} ({col[2]})\n"
            except sqlite3.OperationalError:
                schema_str += f"\n{table}: [Error reading schema]\n"
        
        conn.close()
        
        return jsonify({
            "file": db_path,
            "size": size,
            "modified": modified,
            "tables": tables,
            "schema": schema_str
        })
    except Exception as e:
        return jsonify({"error": f"Database error: {str(e)}"}), 500

@app.route("/api/admin/backup-database", methods=["POST"])
@login_required
def api_backup_database():
    """Create database backup (admin only)"""
    user_id = session['user_id']
    
    if not is_admin(user_id):
        return jsonify({"error": "Admin access required"}), 403
    
    try:
        import shutil
        from datetime import datetime
        from pathlib import Path
        
        # Use same database path as database.py
        main_db_path = Path("../data/aerium.sqlite")
        site_db_path = Path("data/aerium.sqlite")
        
        if main_db_path.exists():
            db_path = str(main_db_path)
        else:
            db_path = str(site_db_path)
        
        if not os.path.exists(db_path):
            return jsonify({"error": "Database file not found"}), 404
        
        # Create backup directory if it doesn't exist
        backup_dir = os.path.dirname(db_path) + "/backups"
        os.makedirs(backup_dir, exist_ok=True)
        
        # Create backup file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"aerium_backup_{timestamp}.sqlite"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        shutil.copy2(db_path, backup_path)
        
        log_audit(user_id, "BACKUP", f"Database backed up to {backup_path}")
        
        # Return the backup file
        return send_file(backup_path, as_attachment=True, download_name=f"morpheus-backup-{timestamp}.sqlite")
    except Exception as e:
        log_audit(user_id, "ERROR", f"Backup failed: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/history/latest/<int:limit>")
def api_history_latest(limit):
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    source_clause, source_params = build_source_filter(db_source)

    db = get_db()
    rows = db.execute(f"""
        SELECT id, ppm, timestamp
        FROM co2_readings
        WHERE {source_clause}
        ORDER BY id DESC
        LIMIT ?
    """, (*source_params, limit)).fetchall()
    db.close()

    # reverse so oldest → newest
    return jsonify([dict(r) for r in reversed(rows)])

@app.route("/api/cleanup", methods=["POST"])
def api_cleanup():
    """Clean up old CO₂ readings (default: 90 days)"""
    days = request.json.get("days", 90) if request.json else 90
    deleted = cleanup_old_data(days)
    return jsonify({"status": "ok", "deleted": deleted, "days": days})

@app.route("/api/reset-state", methods=["POST"])
def api_reset_state():
    """Reset CO₂ generator state"""
    base = request.json.get("base", 600) if request.json else 600
    reset_state(base)
    return jsonify({"status": "ok", "base": base})

@app.route("/api/thresholds", methods=["GET", "POST"])
@login_required
def api_thresholds():
    """Get or update user's CO₂ thresholds"""
    user_id = session['user_id']
    
    if request.method == "GET":
        thresholds = get_user_thresholds(user_id)
        return jsonify({
            "good_level": thresholds['good_level'],
            "warning_level": thresholds['warning_level'],
            "critical_level": thresholds['critical_level']
        }), 200
    
    elif request.method == "POST":
        data = request.json
        if not data:
            return jsonify({"error": "No JSON data provided"}), 400
            
        good = int(data.get('good_level', 800))
        warning = int(data.get('warning_level', 1000))
        critical = int(data.get('critical_level', 1200))
        
        # Validate: good < warning < critical
        if not (good < warning < critical):
            return jsonify({"error": "Invalid threshold order"}), 400
        
        update_user_thresholds(user_id, good, warning, critical)
        log_audit(user_id, "UPDATE", "Thresholds updated")
        
        return jsonify({
            "status": "ok",
            "good_level": good,
            "warning_level": warning,
            "critical_level": critical
        }), 200
    
    return jsonify({"error": "Method not allowed"}), 405

def generate_pdf(html):
    pdf_io = io.BytesIO()

    HTML(
        string=html,
        base_url=os.path.abspath(".")
    ).write_pdf(
        target=pdf_io,
        presentational_hints=True
    )

    pdf_io.seek(0)

    return send_file(
        pdf_io,
        mimetype="application/pdf",
        as_attachment=False,
        download_name="daily_report.pdf"
    )


@app.route("/api/report/daily/pdf")
def export_daily_pdf():
    db_source = resolve_source_param(allow_sim=True, allow_import=True)
    data = get_today_history(db_source)
    settings = load_settings()

    if not data:
        return "No data", 400

    values = [d["ppm"] for d in data]

    avg = round(sum(values) / len(values))
    max_ppm = max(values)
    min_ppm = min(values)

    # ⏱ minutes above bad threshold
    bad_minutes = sum(1 for v in values if v >= settings["bad_threshold"])

    # ✅ EXPOSURE BREAKDOWN (THIS IS YOUR QUESTION)
    good = sum(1 for v in values if v < settings["good_threshold"])
    medium = sum(
        1 for v in values
        if settings["good_threshold"] <= v < settings["bad_threshold"]
    )
    bad = sum(1 for v in values if v >= settings["bad_threshold"])
    total = len(values)

    good_pct = round(good / total * 100)
    medium_pct = round(medium / total * 100)
    bad_pct = round(bad / total * 100)

    with open("static/css/report.css", "r", encoding="utf-8") as f:
        report_css = f.read()

    html = render_template(
        "report_daily.html",
        date=date.today().strftime("%d %B %Y"),
        avg=avg,
        max=max_ppm,
        min=min_ppm,
        bad_minutes=bad_minutes,
        good_pct=good_pct,
        medium_pct=medium_pct,
        bad_pct=bad_pct,
        good_threshold=settings["good_threshold"],
        bad_threshold=settings["bad_threshold"],
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        report_css=report_css
    )

    return generate_pdf(html)


@app.route("/healthz")
def healthz():
    db = get_db()
    latest = db.execute("SELECT ppm, timestamp FROM co2_readings ORDER BY id DESC LIMIT 1").fetchone()
    count = db.execute("SELECT COUNT(*) AS c FROM co2_readings").fetchone()["c"]
    settings = load_settings()
    db.close()

    return jsonify({
        "status": "ok",
        "analysis_running": settings.get("analysis_running", True),
        "rows": count,
        "latest_ppm": latest["ppm"] if latest else None,
        "latest_timestamp": latest["timestamp"] if latest else None,
    })


@app.route("/metrics")
def metrics():
    db = get_db()
    latest = db.execute("SELECT ppm, timestamp FROM co2_readings ORDER BY id DESC LIMIT 1").fetchone()
    count = db.execute("SELECT COUNT(*) AS c FROM co2_readings").fetchone()["c"]
    settings = load_settings()
    db.close()

    payload = {
        "rows": count,
        "analysis_running": settings.get("analysis_running", True),
        "good_threshold": settings.get("good_threshold"),
        "bad_threshold": settings.get("bad_threshold"),
        "update_speed": settings.get("update_speed"),
        "overview_update_speed": settings.get("overview_update_speed"),
        "latest_ppm": latest["ppm"] if latest else None,
        "latest_timestamp": latest["timestamp"] if latest else None,
    }

    return jsonify(payload)


# ================================================================================
#                    HARDWARE SIMULATION CONTROL (TESTING)
# ================================================================================

@app.route("/api/simulator/scenario/<scenario_name>", methods=['POST'])
def set_simulation_scenario(scenario_name):
    """Set simulation scenario for testing"""
    if not is_admin(session.get('user_id')):
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    
    duration = request.json.get('duration', 0) if request.json else 0
    valid_scenarios = ['normal', 'office_hours', 'sleep', 'ventilation', 'anomaly']
    
    if scenario_name not in valid_scenarios:
        return jsonify({'success': False, 'error': f'Invalid scenario. Must be one of: {valid_scenarios}'}), 400
    
    result = set_scenario(scenario_name, duration)
    
    if result:
        return jsonify({
            'success': True,
            'message': f'Scenario set to {scenario_name}',
            'duration': duration,
            'info': get_scenario_info()
        })
    return jsonify({'success': False, 'error': 'Failed to set scenario'}), 400


@app.route("/api/simulator/status", methods=['GET'])
@limiter.exempt
def get_simulator_status():
    """Get current simulator status"""
    info = get_scenario_info()
    return jsonify({
        'success': True,
        'simulator': {
            'scenario': info['scenario'],
            'co2': info['co2'],
            'temperature': info['temp'],
            'humidity': info['humidity'],
            'timer': info['timer'],
            'duration': info['duration'],
            'paused': info.get('paused', False)
        }
    })


@app.route("/api/simulator/pause", methods=['POST'])
def pause_simulator():
    """Pause or resume the simulator progression."""
    desired = False
    if request.json is not None:
        desired = bool(request.json.get('paused', False))

    set_paused(desired)
    info = get_scenario_info()
    return jsonify({
        'success': True,
        'paused': desired,
        'simulator': {
            'scenario': info['scenario'],
            'co2': info['co2'],
            'temperature': info['temp'],
            'humidity': info['humidity'],
            'timer': info['timer'],
            'duration': info['duration'],
            'paused': info.get('paused', desired)
        }
    })


@app.route("/api/simulator/reset", methods=['POST'])
def reset_simulator():
    """Reset simulator to initial state"""
    if not is_admin(session.get('user_id')):
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    
    base_co2 = request.json.get('base_co2', 600) if request.json else 600
    scenario = request.json.get('scenario', 'normal') if request.json else 'normal'
    
    reset_state(base_co2, scenario)
    
    return jsonify({
        'success': True,
        'message': 'Simulator reset',
        'info': get_scenario_info()
    })


# ================================================================================
#                          WEBSOCKET HANDLERS
# ================================================================================

# Global state for WebSocket broadcasting
broadcast_thread = None
broadcast_running = False

@socketio.on('connect')
def handle_connect():
    """Handle client connection to WebSocket"""
    print(f"Client connected")
    emit('status', {'data': 'Connected to Morpheus CO₂ Monitor'})
    
    # Send current settings to client
    settings = load_settings()
    emit('settings_update', settings)

@socketio.on('disconnect')
def handle_disconnect():
    """Handle client disconnection"""
    print(f"Client disconnected")

@socketio.on('request_data')
def handle_request_data():
    """Handle request for latest CO₂ data"""
    _, payload = build_live_payload()
    emit('co2_update', payload)

@socketio.on('settings_change')
def handle_settings_change(data):
    """Handle settings update and broadcast to all clients"""
    save_settings(data)
    # Broadcast to all clients
    socketio.emit('settings_update', data)

def broadcast_co2_loop():
    """Background thread that broadcasts CO₂ readings to all connected clients"""
    global broadcast_running
    broadcast_running = True
    last_ppm = None
    last_state = None  # 'running', 'paused', 'no_sensor'
    
    while broadcast_running:
        settings = load_settings()
        update_delay = settings.get("update_speed", 0.5)

        settings, payload = build_live_payload(settings)
        current_state = "running" if payload.get("analysis_running") else payload.get("reason", "paused")

        if not payload.get("analysis_running"):
            if last_state != current_state:
                socketio.emit('co2_update', {
                    'analysis_running': False,
                    'reason': payload.get('reason'),
                    'ppm': None,
                    'timestamp': datetime.now(UTC).isoformat()
                }, to=None)
                last_state = current_state
            time.sleep(update_delay)
            continue

        socketio.emit('co2_update', payload, to=None)
        last_ppm = payload.get('ppm')
        last_state = current_state

        time.sleep(update_delay)

def start_broadcast_thread():
    """Start the background broadcast thread"""
    global broadcast_thread, broadcast_running
    if broadcast_thread is None or not broadcast_thread.is_alive():
        broadcast_running = True
        broadcast_thread = threading.Thread(target=broadcast_co2_loop, daemon=True)
        broadcast_thread.start()
        print("[OK] WebSocket broadcast thread started")

def stop_broadcast_thread():
    """Stop the background broadcast thread"""
    global broadcast_running
    broadcast_running = False
    print("[OK] WebSocket broadcast thread stopped")




if __name__ == "__main__":
    # Database migrations
    try:
        db = get_db()
        
        # Add temperature and humidity columns to co2_readings if missing
        try:
            db.execute("ALTER TABLE co2_readings ADD COLUMN temperature REAL")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE co2_readings ADD COLUMN humidity REAL")
        except Exception:
            pass
        
        # Migrate audit_logs table to add missing columns
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN user_id INTEGER")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN username TEXT")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN entity_type TEXT")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN entity_id TEXT")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN details TEXT")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN status TEXT")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE audit_logs ADD COLUMN severity TEXT")
        except Exception:
            pass
        
        db.commit()
        db.close()
    except Exception as e:
        print(f"Migration error: {e}")

    start_broadcast_thread()
    socketio.run(app, debug=True, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True)