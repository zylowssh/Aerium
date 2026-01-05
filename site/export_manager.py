"""
Export Module - PDF and Excel report generation
Handles data export with formatting, charts, and styling
"""

import io
import csv
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import json

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataExporter:
    """Export CO₂ data in various formats"""
    
    def __init__(self):
        self.weasyprint_available = WEASYPRINT_AVAILABLE
        self.openpyxl_available = OPENPYXL_AVAILABLE
    
    def export_to_csv(self, data: List[Dict], filename: str = "export") -> io.BytesIO:
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries to export
            filename: Name for the file (without extension)
        
        Returns:
            BytesIO object containing CSV data
        """
        output = io.StringIO()
        
        if not data:
            return io.BytesIO()
        
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
        
        output.seek(0)
        bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
        return bytes_output
    
    def export_to_excel(self, data: List[Dict], title: str = "CO₂ Report", 
                       charts_data: Optional[Dict] = None) -> Optional[io.BytesIO]:
        """
        Export data to Excel with formatting and optional charts
        
        Args:
            data: List of dictionaries to export
            title: Report title
            charts_data: Optional dictionary with chart configurations
        
        Returns:
            BytesIO object containing Excel data or None if openpyxl unavailable
        """
        if not self.openpyxl_available:
            return None
        
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, BarChart
        from openpyxl.chart.reference import Reference
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Add timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Add data with headers
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Add data rows
            for row_idx, row_data in enumerate(data, 5):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = row_data.get(header)
                    
                    # Format numbers
                    if isinstance(cell.value, (int, float)) and col_idx > 1:
                        cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_to_pdf(self, html_content: str, filename: str = "export") -> Optional[io.BytesIO]:
        """
        Export HTML content to PDF
        
        Args:
            html_content: HTML string to convert to PDF
            filename: Name for the file (without extension)
        
        Returns:
            BytesIO object containing PDF data or None if WeasyPrint unavailable
        """
        if not self.weasyprint_available:
            return None
        
        try:
            html = HTML(string=html_content)
            pdf_bytes = html.write_pdf()
            return io.BytesIO(pdf_bytes)
        except Exception as e:
            print(f"PDF generation error: {e}")
            return None
    
    def generate_report_html(self, title: str, summary: Dict, 
                            data: List[Dict], charts: Optional[Dict] = None) -> str:
        """
        Generate HTML report with styling
        
        Args:
            title: Report title
            summary: Dictionary of summary statistics
            data: Data rows to display
            charts: Optional chart configurations
        
        Returns:
            HTML string
        """
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 40px;
                    color: #333;
                    background: white;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 30px;
                    border-bottom: 3px solid #366092;
                    padding-bottom: 20px;
                }}
                .header h1 {{
                    margin: 0;
                    color: #366092;
                }}
                .timestamp {{
                    color: #666;
                    font-size: 12px;
                }}
                .summary {{
                    display: grid;
                    grid-template-columns: repeat(4, 1fr);
                    gap: 20px;
                    margin: 20px 0;
                }}
                .summary-card {{
                    background: #f5f5f5;
                    padding: 15px;
                    border-radius: 8px;
                    text-align: center;
                    border-left: 4px solid #366092;
                }}
                .summary-card h3 {{
                    margin: 0 0 10px 0;
                    font-size: 14px;
                    color: #666;
                }}
                .summary-card .value {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #366092;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th {{
                    background: #366092;
                    color: white;
                    padding: 12px;
                    text-align: left;
                    font-weight: bold;
                }}
                td {{
                    padding: 10px;
                    border-bottom: 1px solid #ddd;
                }}
                tr:nth-child(even) {{
                    background: #f9f9f9;
                }}
                .footer {{
                    text-align: center;
                    margin-top: 30px;
                    color: #999;
                    font-size: 12px;
                }}
                @media print {{
                    body {{ margin: 0; }}
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <div class="timestamp">Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
            </div>
        """
        
        # Add summary cards
        if summary:
            html += '<div class="summary">'
            for key, value in summary.items():
                html += f"""
                <div class="summary-card">
                    <h3>{key}</h3>
                    <div class="value">{value}</div>
                </div>
                """
            html += '</div>'
        
        # Add data table
        if data:
            html += '<table><thead><tr>'
            headers = list(data[0].keys())
            for header in headers:
                html += f'<th>{header}</th>'
            html += '</tr></thead><tbody>'
            
            for row in data[:100]:  # Limit to 100 rows for PDF
                html += '<tr>'
                for header in headers:
                    value = row.get(header, '')
                    html += f'<td>{value}</td>'
                html += '</tr>'
            
            html += '</tbody></table>'
        
        html += """
            <div class="footer">
                <p>This is an automated report. For questions, contact your administrator.</p>
            </div>
        </body>
        </html>
        """
        
        return html


class ScheduledExporter:
    """Handle scheduled exports of data"""
    
    def __init__(self, db_connection):
        self.db = db_connection
        self.exporter = DataExporter()
    
    def create_export_schedule(self, user_id: int, export_type: str, 
                              frequency: str, format_type: str) -> bool:
        """
        Schedule periodic exports for a user
        
        Args:
            user_id: User ID
            export_type: Type of data to export (e.g., 'readings', 'analytics')
            frequency: Export frequency ('daily', 'weekly', 'monthly')
            format_type: Export format ('csv', 'excel', 'pdf')
        
        Returns:
            True if schedule created successfully
        """
        try:
            from database import create_scheduled_export
            create_scheduled_export(user_id, export_type, frequency, format_type)
            return True
        except Exception as e:
            print(f"Error creating export schedule: {e}")
            return False
    
    def process_due_exports(self) -> List[Tuple[int, str, str]]:
        """
        Process all due exports and return list of (user_id, file_path, format)
        
        Returns:
            List of tuples containing export information
        """
        results = []
        try:
            from database import get_due_exports, update_export_timestamp
            due_exports = get_due_exports()
            
            for export in due_exports:
                # Generate export data
                # This would be integrated with your data retrieval
                update_export_timestamp(export['id'])
                results.append((export['user_id'], f"export_{export['id']}", export['format']))
        
        except Exception as e:
            print(f"Error processing exports: {e}")
        
        return results
